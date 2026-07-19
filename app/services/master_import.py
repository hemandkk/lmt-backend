"""Parse CSV / Excel rows for masters import."""

from __future__ import annotations

import csv
import io
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return (
        text.replace(" ", "")
        .replace("_", "")
        .replace("-", "")
    )


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_bool(value: Any, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "active"}:
        return True
    if text in {"0", "false", "no", "n", "inactive"}:
        return False
    return default


def _parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


async def read_tabular_rows(file: UploadFile) -> list[dict[str, Any]]:
    """
    Read CSV or XLSX into a list of dicts keyed by normalized headers.
    First row must be headers.
    """
    filename = (file.filename or "").lower()
    raw = await file.read()
    if not raw:
        raise ValueError("Uploaded file is empty.")

    if filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        workbook = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise ValueError("Spreadsheet has no rows.") from exc
        headers = [_norm_header(h) for h in header_row]
        if not any(headers):
            raise ValueError("Spreadsheet header row is empty.")
        result: list[dict[str, Any]] = []
        for row in rows_iter:
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            item: dict[str, Any] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                item[header] = row[idx] if idx < len(row) else None
            result.append(item)
        return result

    # Default: treat as CSV (also .txt)
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV header row is missing.")
    result = []
    for row in reader:
        if not row or all(not str(v or "").strip() for v in row.values()):
            continue
        item = {_norm_header(k): v for k, v in row.items() if k is not None}
        result.append(item)
    return result


def map_course_row(row: dict[str, Any]) -> dict[str, Any]:
    name = _cell_str(
        row.get("name")
        or row.get("coursename")
        or row.get("course")
    )
    if not name:
        raise ValueError("name is required")

    active_raw = row.get("active")
    if active_raw is None:
        active_raw = row.get("isactive")

    return {
        "name": name,
        "course_code": _cell_str(
            row.get("coursecode") or row.get("code")
        ),
        "specialization": _cell_str(row.get("specialization")),
        "duration": _cell_str(row.get("duration")),
        "fees": _parse_int(row.get("fees")),
        "description": _cell_str(row.get("description")),
        "is_active": _parse_bool(active_raw, default=True),
    }


def map_specialization_row(row: dict[str, Any]) -> dict[str, Any]:
    name = _cell_str(
        row.get("name")
        or row.get("specialization")
        or row.get("specializationname")
    )
    if not name:
        raise ValueError("name is required")

    active_raw = row.get("active")
    if active_raw is None:
        active_raw = row.get("isactive")

    return {
        "name": name,
        "specialization_code": _cell_str(
            row.get("specializationcode") or row.get("code")
        ),
        "description": _cell_str(row.get("description")),
        "is_active": _parse_bool(active_raw, default=True),
    }
